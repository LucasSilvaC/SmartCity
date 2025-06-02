from rest_framework import status, filters
from rest_framework.decorators import api_view, permission_classes
from rest_framework.parsers import MultiPartParser
from rest_framework.response import Response
from rest_framework.generics import ListCreateAPIView, RetrieveUpdateDestroyAPIView
from rest_framework.permissions import IsAuthenticated
from rest_framework.views import APIView
from django_filters.rest_framework import DjangoFilterBackend
from django.http import HttpResponse
from openpyxl import load_workbook, Workbook
from openpyxl.utils import get_column_letter
from datetime import datetime
from rest_framework_simplejwt.tokens import RefreshToken

from .models import (
    User,
    Sensor,
    Ambiente,
    Historico
)
from .serializers import (
    SensorSerializer,
    AmbienteSerializer,
    HistoricoSerializer
)

# Criação de superusuário
class CriarUsuarioAPIView(APIView):
    def post(self, request):
        username = request.data.get('username')
        password = request.data.get('password')
        if not username or not password:
            return Response({'erro': 'Usuário e senha são obrigatórios'}, status=status.HTTP_400_BAD_REQUEST)

        if User.objects.filter(username=username).exists():
            return Response({'erro': 'Usuário já existe'}, status=status.HTTP_400_BAD_REQUEST)

        user = User.objects.create_superuser(username=username, password=password)
        refresh = RefreshToken.for_user(user)
        return Response({
            'mensagem': 'Superusuário criado com sucesso!',
            'access_token': str(refresh.access_token)
        }, status=status.HTTP_201_CREATED)

# Reset password
class ResetPasswordAPIView(APIView):
    def post(self, request):
        username = request.data.get('username')
        new_password = request.data.get('new_password')

        if not username or not new_password:
            return Response({'erro': 'Usuário e nova senha são obrigatórios'}, status=status.HTTP_400_BAD_REQUEST)

        try:
            user = User.objects.get(username=username)
            user.set_password(new_password)
            user.save()
            return Response({'mensagem': 'Senha redefinida com sucesso!'}, status=status.HTTP_200_OK)
        except User.DoesNotExist:
            return Response({'erro': 'Usuário não encontrado'}, status=status.HTTP_404_NOT_FOUND)

# Ambientes
class AmbienteListCreateView(ListCreateAPIView):
    queryset = Ambiente.objects.all().order_by('sig')
    serializer_class = AmbienteSerializer
    permission_classes = [IsAuthenticated]
    filter_backends = [DjangoFilterBackend, filters.SearchFilter]
    filterset_fields = ['sig']
    search_fields = ['sig']

class AmbienteDetailView(RetrieveUpdateDestroyAPIView):
    queryset = Ambiente.objects.all()
    serializer_class = AmbienteSerializer
    permission_classes = [IsAuthenticated]

class UploadXLSXViewAmbiente(APIView):
    permission_classes = [IsAuthenticated]
    parser_classes = [MultiPartParser]

    def post(self, request, *args, **kwargs):
        file_obj = request.FILES.get('file')

        if not file_obj:
            return Response({'erro': 'Arquivo não enviado'}, status=400)

        wb = load_workbook(filename=file_obj)
        ws = wb.active

        for i, row in enumerate(ws.iter_rows(min_row=2, values_only=True)):
            sig = row[0]
            descricao = row[1]
            ni = row[2]
            responsavel = row[3]

            if not ni:
                print(f"[Linha {i+2}] Erro: ni vazio. Dados: {row}")
                continue

            Ambiente.objects.create(
                sig=sig,
                descricao=descricao,
                ni=ni,
                responsavel=responsavel,
            )

        return Response({'mensagem': 'Dados importados com sucesso!'}, status=status.HTTP_201_CREATED)

@api_view(['GET'])
@permission_classes([IsAuthenticated])
def exportar_xlsx_ambientes(request):
    wb = Workbook()
    ws = wb.active
    ws.title = "Ambientes"

    headers = ['sig', 'descricao', 'ni', 'responsavel']
    ws.append(headers)

    for ambiente in Ambiente.objects.all():
        ws.append([
            ambiente.sig,
            ambiente.descricao,
            ambiente.ni,
            ambiente.responsavel,
        ])

    for col in ws.columns:
        max_length = max(len(str(cell.value)) if cell.value else 0 for cell in col)
        ws.column_dimensions[get_column_letter(col[0].column)].width = max_length + 5

    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = 'attachment; filename="ambientes.xlsx"'
    wb.save(response)
    return response


# Sensores
class SensorListCreateView(ListCreateAPIView):
    queryset = Sensor.objects.all()
    serializer_class = SensorSerializer
    permission_classes = [IsAuthenticated]
    filter_backends = [DjangoFilterBackend, filters.SearchFilter]
    filterset_fields = ['id', 'sensor', 'status']
    search_fields = ['sensor']

class SensorDetailView(RetrieveUpdateDestroyAPIView):
    queryset = Sensor.objects.all()
    serializer_class = SensorSerializer
    permission_classes = [IsAuthenticated]

def str_to_bool(val):
    if isinstance(val, str):
        texto = val.strip().lower()
        return texto in ("true", "1", "sim", "ativo", "yes")
    return bool(val)

class UploadXLSXViewSensor(APIView):
    permission_classes = [IsAuthenticated]
    parser_classes = [MultiPartParser]

    def post(self, request, *args, **kwargs):
        file_obj = request.FILES.get('file')
        if not file_obj:
            return Response({'erro': 'Arquivo não enviado'}, status=400)

        try:
            wb = load_workbook(filename=file_obj, data_only=True)
            ws = wb.active
        except Exception as e:
            return Response({'erro': f'Falha ao abrir o Excel: {e}'}, status=400)

        header_row = next(ws.iter_rows(min_row=1, max_row=1, values_only=True))
        col_idx = {}
        for idx, cell_value in enumerate(header_row):
            if cell_value is None:
                continue
            nome = str(cell_value).strip()
            col_idx[nome] = idx

        obrigatorias = ['sensor', 'mac_address', 'latitude', 'longitude', 'status']
        faltantes = [c for c in obrigatorias if c not in col_idx]
        if faltantes:
            return Response(
                {'erro': f'Colunas obrigatórias faltando no Excel: {faltantes}'},
                status=400
            )

        erros = []
        criados = 0
        atualizados = 0

        for linha_idx, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
            tipo_sensor_raw = row[col_idx['sensor']]
            mac_address_raw = row[col_idx['mac_address']]
            latitude_raw = row[col_idx['latitude']]
            longitude_raw = row[col_idx['longitude']]
            status_raw = row[col_idx['status']]

            if tipo_sensor_raw is None or str(tipo_sensor_raw).strip() == "":
                erros.append(f"[Linha {linha_idx}] Erro: coluna 'sensor' vazia. Dados: {row}")
                continue
            if mac_address_raw is None or str(mac_address_raw).strip() == "":
                erros.append(f"[Linha {linha_idx}] Erro: coluna 'mac_address' vazia. Dados: {row}")
                continue

            tipo_sensor = str(tipo_sensor_raw).strip().lower()
            mac_address = str(mac_address_raw).strip()

            try:
                latitude = float(latitude_raw)
                longitude = float(longitude_raw)
            except Exception:
                erros.append(f"[Linha {linha_idx}] Erro: latitude/longitude inválidos. Dados: {row}")
                continue

            status_val = str_to_bool(status_raw)

            tipos_validos = {choice[0] for choice in Sensor.TIPO}
            if tipo_sensor not in tipos_validos:
                erros.append(f"[Linha {linha_idx}] Erro: tipo '{tipo_sensor}' não está entre {tipos_validos}.")
                continue

            identificador_unico = f"{tipo_sensor}_{mac_address}"

            conflito_sensor = Sensor.objects.filter(sensor=identificador_unico).exclude(mac_address=mac_address).exists()
            if conflito_sensor:
                erros.append(
                    f"[Linha {linha_idx}] Erro: Já existe outro Sensor com nome '{identificador_unico}' e MAC diferente."
                )
                continue

            try:
                obj, created = Sensor.objects.update_or_create(
                    mac_address=mac_address,
                    defaults={
                        'sensor': identificador_unico,
                        'unidade_med': tipo_sensor,
                        'latitude': latitude,
                        'longitude': longitude,
                        'status': status_val,
                    }
                )
                if created:
                    criados += 1
                else:
                    atualizados += 1
            except Exception as e:
                erros.append(f"[Linha {linha_idx}] Erro ao salvar no banco: {e}")
                continue

        resposta = {
            'total_linhas_lidas': ws.max_row - 1,
            'criados': criados,
            'atualizados': atualizados
        }
        if erros:
            resposta['erros'] = erros
            return Response(resposta, status=status.HTTP_207_MULTI_STATUS)

        return Response(resposta, status=status.HTTP_201_CREATED)

@api_view(['GET'])
@permission_classes([IsAuthenticated])
def exportar_xlsx_sensores(request):
    wb = Workbook()
    ws = wb.active
    ws.title = "Sensores"

    headers = ['sensor', 'mac_address', 'unidade_med', 'latitude', 'longitude', 'status']
    ws.append(headers)

    for sensor in Sensor.objects.all():
        ws.append([
            sensor.sensor,
            sensor.mac_address,
            sensor.unidade_med,
            sensor.latitude,
            sensor.longitude,
            sensor.status,
        ])

    for col in ws.columns:
        max_length = max(len(str(cell.value)) if cell.value else 0 for cell in col)
        ws.column_dimensions[get_column_letter(col[0].column)].width = max_length + 5

    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = 'attachment; filename="sensores.xlsx"'
    wb.save(response)
    return response


# Histórico
class HistoricoListCreateView(ListCreateAPIView):
    queryset = Historico.objects.all().order_by('-timestamp')
    serializer_class = HistoricoSerializer
    permission_classes = [IsAuthenticated]
    filter_backends = [DjangoFilterBackend, filters.SearchFilter]
    filterset_fields = ['sensor', 'ambiente', 'timestamp']
    search_fields = ['sensor__sensor', 'ambiente__descricao']

    def get_queryset(self):
        queryset = super().get_queryset()
        sensor_id = self.request.query_params.get('sensor')
        ambiente_id = self.request.query_params.get('ambiente')
        data_str = self.request.query_params.get('data')
        hora_str = self.request.query_params.get('hora')

        if sensor_id:
            queryset = queryset.filter(sensor__id=sensor_id)
        if ambiente_id:
            queryset = queryset.filter(ambiente__id=ambiente_id)
        if data_str:
            queryset = queryset.filter(timestamp__date=data_str)
        if hora_str:
            try:
                hora_int = int(hora_str)
                queryset = queryset.filter(timestamp__hour=hora_int)
            except ValueError:
                pass

        return queryset

class HistoricoDetailView(RetrieveUpdateDestroyAPIView):
    queryset = Historico.objects.all()
    serializer_class = HistoricoSerializer
    permission_classes = [IsAuthenticated]

class UploadXLSXViewHistorico(APIView):
    permission_classes = [IsAuthenticated]
    parser_classes = [MultiPartParser]

    def post(self, request, *args, **kwargs):
        file_obj = request.FILES.get('file')

        if not file_obj:
            return Response({'erro': 'Arquivo não enviado'}, status=400)

        wb = load_workbook(filename=file_obj)
        ws = wb.active

        erros = []
        for i, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
            valor = row[0]
            ambiente_raw = row[1]
            sensor_raw = row[2]
            timestamp = row[3]

            if not valor or not timestamp or not ambiente_raw or not sensor_raw:
                erros.append(f"[Linha {i}] Dados incompletos. Dados: {row}")
                continue

            try:
                if isinstance(ambiente_raw, float):
                    ambiente_id = int(ambiente_raw)
                elif isinstance(ambiente_raw, str):
                    ambiente_id = int(ambiente_raw.strip())
                else:
                    ambiente_id = int(ambiente_raw)
            except (ValueError, TypeError):
                erros.append(f"[Linha {i}] Ambiente ID inválido: {ambiente_raw}")
                continue

            try:
                if isinstance(sensor_raw, float):
                    sensor_id = int(sensor_raw)
                elif isinstance(sensor_raw, str):
                    sensor_id = int(sensor_raw.strip())
                else:
                    sensor_id = int(sensor_raw)
            except (ValueError, TypeError):
                erros.append(f"[Linha {i}] Sensor ID inválido: {sensor_raw}")
                continue

            try:
                sensor = Sensor.objects.get(id=sensor_id)
            except Sensor.DoesNotExist:
                erros.append(f"[Linha {i}] Sensor com ID {sensor_id} não encontrado.")
                continue

            try:
                ambiente = Ambiente.objects.get(id=ambiente_id)
            except Ambiente.DoesNotExist:
                erros.append(f"[Linha {i}] Ambiente com ID {ambiente_id} não encontrado.")
                continue

            Historico.objects.create(
                valor=valor,
                timestamp=timestamp,
                ambiente=ambiente,
                sensor=sensor,
            )

        if erros:
            return Response({'mensagem': 'Importação concluída com erros', 'erros': erros}, status=status.HTTP_207_MULTI_STATUS)

        return Response({'mensagem': 'Dados importados com sucesso!'}, status=status.HTTP_201_CREATED)

@api_view(['GET'])
@permission_classes([IsAuthenticated])
def exportar_historico_excel(request):
    hora = request.GET.get('hora')
    data = request.GET.get('data')

    if not hora or not data:
        return HttpResponse("Parâmetros 'hora' e 'data' são obrigatórios.", status=400)

    try:
        data_obj = datetime.strptime(data, "%Y-%m-%d").date()
        hora_int = int(hora)
    except ValueError:
        return HttpResponse("Formato inválido para data ou hora.", status=400)

    historico = Historico.objects.filter(timestamp__date=data_obj, timestamp__hour=hora_int)

    wb = Workbook()
    ws = wb.active
    ws.title = "Histórico"

    headers = ['ID', 'Sensor', 'Ambiente', 'Valor', 'Timestamp']
    ws.append(headers)

    for h in historico:
        ws.append([
            h.id,
            str(h.sensor),
            str(h.ambiente),
            h.valor,
            h.timestamp.strftime('%Y-%m-%d %H:%M:%S')
        ])

    for col in ws.columns:
        max_length = max(len(str(cell.value)) for cell in col)
        ws.column_dimensions[get_column_letter(col[0].column)].width = max_length + 2

    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = f'attachment; filename=historico_{data}_{hora}h.xlsx'
    wb.save(response)
    return response